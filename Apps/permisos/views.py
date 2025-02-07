from django.utils.timezone import is_aware
from .models import registroPermisos  # Importa tu modelo
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
import openpyxl
import datetime
from .models import registroPermisos, Colaboradores, Departamento, tiposPermiso
from django.db.models import F
from django.shortcuts import render
from django.http import JsonResponse
from django.utils import timezone
from .utils import enviar_correo_permiso
from django.utils.timezone import now
from . models import Empresas, Sucursal
from django.shortcuts import get_object_or_404
from openpyxl.worksheet.table import Table, TableStyleInfo


from django.contrib.auth.decorators import login_required, user_passes_test, permission_required
from .models import registroPermisos

# Create your views here.

# LISTA DE DEPARTAMENTOS
def lista_departamentos_view(request):
    departamentos = Departamento.objects.filter(
        estado='ACTIVO').values('id', 'nombre_departamento')
    return JsonResponse(list(departamentos), safe=False)

def lista_empresas_view(request):
    empresas = Empresas.objects.filter(
        estado='ACTIVO').values('id', 'nombre_empresa')
    return JsonResponse(list(empresas), safe=False)

def lista_sucursales_view(request):
    sucursales = Sucursal.objects.filter(
        estado='ACTIVO').values('id', 'nombre_sucursal')
    return JsonResponse(list(sucursales), safe=False)

def colaboradores_por_departamento_view(request, departamento_id):
    colaboradores = Colaboradores.objects.filter(
        departamento_id=departamento_id, estado='ACTIVO'
    ).values(
        'id', 'nombrecolaborador', 'codigocolaborador', 'empresa_id', 'sucursal_id'
    )

    return JsonResponse(list(colaboradores), safe=False)


def verificar_colaborador_view(request):
    empresa_id = request.GET.get("empresa")
    sucursal_id = request.GET.get("sucursal")
    nombre_colaborador = request.GET.get("nombre")

    if not (empresa_id and sucursal_id and nombre_colaborador):
        return JsonResponse({"error": "Faltan parámetros"}, status=400)

    try:
        existe = Colaboradores.objects.filter(
            empresa_id=empresa_id,
            sucursal_id=sucursal_id,
            nombrecolaborador=nombre_colaborador,
            estado='ACTIVO'
        ).exists()

        return JsonResponse({"existe": existe})
    except Exception as e:
        return JsonResponse({"error": f"Error en la validación: {e}"}, status=400)


def cargar_colaboradores(request, jefe_id):
    colaboradores = Colaboradores.objects.filter(
        jefe_id=jefe_id, estado='ACTIVO')
    data = {
        "colaboradores": list(colaboradores.values('id', 'nombrecolaborador')),
    }
    return JsonResponse(data)

# SOLICITUD DE PERMISOS
@login_required
@permission_required('permisos.add_registropermisos', raise_exception=True)
def permisos_registro_view(request):
    if request.method == "POST":
        
        try:
            # Captura los datos del formulario
            id_empresa = request.POST.get("id_empresa")
            id_sucursal = request.POST.get("id_sucursal")
            id_departamento = request.POST.get("id_departamento")
            colaborador_nombre = request.POST.get("colaborador")
            id_tipo_permiso = request.POST.get("id_tipo_permiso")
            permiso_de = request.POST.get("permiso_de")
            fecha_inicio = request.POST.get("fecha_inicio")
            fecha_fin = request.POST.get("fecha_fin")
            motivo = request.POST.get("motivo")
            comprobante = request.FILES.get("comprobante")
            creado_por = request.user.username
            

            if not id_tipo_permiso:
                return JsonResponse({"status": "Error", "message": "Debe de seleccionar un tipo de permiso."}, status=400)

            # Busca el ID del colaborador a partir del nombre
            try:
                colaborador_obj = Colaboradores.objects.get(
                    nombrecolaborador=colaborador_nombre)
            except Colaboradores.DoesNotExist:
                return JsonResponse({"message": "El colaborador especificado no existe."}, status=400)

            # Verificar si el colaborador ya tiene un permiso activo
            solicitudes_existentes = registroPermisos.objects.filter(
                codigocolaborador=colaborador_obj,
                estado_inicial__in=["Pendiente", "Pre-Aprobado", "Aprobado"]
            )

            if solicitudes_existentes.exists():
                return JsonResponse({"status": "Error",
                                     "message": "El colaborador ya tiene una solicitud de permiso activa. No puede enviar otra."
                                     }, status=400)

            # Crea un nuevo registro
            nuevo_permiso = registroPermisos.objects.create(
                id_empresa_id=id_empresa,
                id_sucursal_id=id_sucursal,
                id_departamento_id=id_departamento,
                codigocolaborador=colaborador_obj,
                colaborador=colaborador_nombre,
                id_tipo_permiso_id=id_tipo_permiso,
                permiso_de=permiso_de,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                motivo=motivo,
                comprobante=comprobante,
                estado_inicial="Pendiente",
                estado_final="Pendiente",
                creado_por= creado_por,
                fecha_creacion=now()
            )
            # Enviar el correo con los detalles del permiso
            enviar_correo_permiso(nuevo_permiso)

            # return redirect('permisos_solicitud_exito')
            return JsonResponse({"status": "Success", "message": "Solicitud de permiso enviada correctamente."}, status=200)
        except Exception as e:
            # print(f"Error al guardar el permiso: {e}")
            return JsonResponse({"status": "Error", "message": f"Error al guardar el la solicitud: {e}"}, status=400)

    return render(request, "solicitud.html")
    # return JsonResponse({"status": "Error", "message": "Método no permitido"}, status=405)


def verificar_solicitud_activa_view(request):
    if request.method == 'POST':
        import json
        body = json.loads(request.body)  # Parsear el cuerpo de la solicitud
        # Obtener el nombre del empleado
        nombre_empleado = body.get('nombreEmpleado')

        if not nombre_empleado:
            return JsonResponse({'error': 'El nombre del empleado es requerido.'}, status=400)

        # Verificar si el empleado tiene una solicitud activa
        tiene_solicitud_activa = registroPermisos.objects.filter(
            codigocolaborador__nombrecolaborador=nombre_empleado,
            fecha_fin__gte=timezone.now()  # Asegurarse de que la solicitud no haya finalizado
        ).exists()

        return JsonResponse({'activa': tiene_solicitud_activa})
    return JsonResponse({'error': 'Método no permitido'}, status=405)


def exito_view(request):
    return render(request, 'permisos/exito.html')


@login_required
@permission_required('permisos.view_registropermisos', raise_exception=True)
def permisos_gestion_view(request):
    # Consulta para obtener los datos relacionados
    permisos_gestion = registroPermisos.objects.select_related(
        'id_departamento', 'id_tipo_permiso', 'codigocolaborador', 'id_empresa', 'id_sucursal'
    ).annotate(
        nombre_departamento=F('id_departamento__nombre_departamento'),
        nombre_tipo_permiso=F('id_tipo_permiso__nombre_permiso'),
        nombre_colaborador=F('codigocolaborador__nombrecolaborador'),
        empresa=F('id_empresa__nombre_empresa'),
        sucursal=F('id_sucursal__nombre_sucursal'),
    )

    return render(request, 'permisos_gestion.html', {'permisos_gestion': permisos_gestion})


 # HISTORIAL DE PERMISOS
@login_required
# @user_passes_test(es_rrhh)
def permisos_historial_view(request):
    permisos = registroPermisos.objects.select_related(
        'codigocolaborador', 'id_tipo_permiso', 'id_empresa', 'id_sucursal', 'id_departamento'
    ).annotate(
        nombre_colaborador=F('codigocolaborador__nombrecolaborador'),
        tipo_permiso=F('id_tipo_permiso__nombre_permiso'),
        nombre_empresa=F('id_empresa__nombre_empresa'),
        nombre_sucursal=F('id_sucursal__nombre_sucursal'),
        nombre_departamento=F('id_departamento__nombre_departamento'),
    )
    return render(request, 'permisos_historial.html', {'permisos': permisos})

def permisos_aprobados_view(request):
    permisos = registroPermisos.objects.select_related(
        'codigocolaborador', 'id_tipo_permiso', 'id_empresa', 'id_sucursal', 'id_departamento'
    ).filter(
        estado_inicial="Pre-Aprobado", estado_final="Pendiente"
    ).annotate(
        nombre_colaborador=F('codigocolaborador__nombrecolaborador'),
        tipo_permiso=F('id_tipo_permiso__nombre_permiso'),
        nombre_empresa=F('id_empresa__nombre_empresa'),
        nombre_sucursal=F('id_sucursal__nombre_sucursal'),
        nombre_departamento=F('id_departamento__nombre_departamento'),
    )

    return render(request, 'permisos_pendientes.html', {'permisos': permisos} )

@login_required
def cargar_comprobantes_view(request):
    return render(request, 'comprobantes.html')

def exportar_permisos_excel(request):
    # Crear un nuevo libro de trabajo y una hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Permisos de Empleados"

    # Definir estilo de encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")  # Color azul para encabezado
    border_style = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )

    # Definir las columnas
    columnas = [
        "ID Permiso", "Fecha Creación", "Colaborador", "Tipo de Permiso", "Permiso de",
        "Fecha Inicio", "Fecha Fin", "Motivo", "Estado Inicial",
        "Estado Final", "Creado Por", "Modificado Por", "Fecha Modificación",
        "Código Colaborador", "Empresa", "Sucursal", "Departamento"
    ]

    ws.append(columnas)

    # Aplicar estilos a la cabecera
    for col_num, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Obtener datos de la base de datos
    permisos = registroPermisos.objects.all().values_list(
        "id_permiso", "fecha_creacion", "colaborador", "id_tipo_permiso__nombre_permiso", "permiso_de",
        "fecha_inicio", "fecha_fin", "motivo", "estado_inicial",
        "estado_final", "creado_por", "modificado_por", "fecha_modificacion",
        "codigocolaborador_id", "id_empresa__nombre_empresa", "id_sucursal__nombre_sucursal",
        "id_departamento__nombre_departamento"
    )

    # Agregar filas con datos, eliminando la zona horaria en las fechas
    for row_index, permiso in enumerate(permisos, start=2):  # Comienza en la fila 2 porque la primera es el encabezado
        permiso = list(permiso)  # Convertir a lista para modificar valores

        # Convertir fechas con zona horaria a naive datetime
        for i in [1, 5, 6, 12]:  # Índices donde están las fechas
            if isinstance(permiso[i], (datetime.datetime, datetime.date)):
                if is_aware(permiso[i]):  # Si tiene zona horaria, quitarla
                    permiso[i] = permiso[i].replace(tzinfo=None)

        ws.append(permiso)

        # Aplicar bordes a cada celda
        for col_num in range(1, len(columnas) + 1):
            cell = ws.cell(row=row_index, column=col_num)
            cell.border = border_style
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Crear una tabla con estilo
    tabla = Table(displayName="TablaPermisos", ref=f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}")

    estilo_tabla = TableStyleInfo(
        name="TableStyleMedium9",  
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.tableStyleInfo = estilo_tabla
    ws.add_table(tabla)

    # Ajustar ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Ajuste de ancho

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Permisos_Empleados.xlsx"'
    wb.save(response)
    
    return response

def guardar_politicas(request):
    return render(request, '')
@login_required
# @user_passes_test(es_gestor_o_jefe)
def permisos_pendientes_view(request):
    usuario = request.user

    if hasattr(usuario, "colaborador") and usuario.colaborador.departamento:
        permisos = registroPermisos.objects.filter(
            id_departamento=usuario.colaborador.departamento,
            estado_inicial="PENDIENTE"
        ).select_related("id_departamento", "id_tipo_permiso", "codigocolaborador")
    else:
        permisos = registroPermisos.objects.none()  # Retorna vacío si no hay departamento

    permisos_list = [
        {
            "id": permiso.id_permiso,
            "departamento": permiso.id_departamento.nombre_departamento,
            "colaborador": permiso.codigocolaborador.nombrecolaborador,
            "tipo_permiso": permiso.id_tipo_permiso.nombre_permiso,
            "fecha_inicio": permiso.fecha_inicio.strftime("%d-%m-%Y"),
            "fecha_fin": permiso.fecha_fin.strftime("%d-%m-%Y"),
            "motivo": permiso.motivo,
            "comprobante": permiso.comprobante.url if permiso.comprobante else None,
            "estado_inicial": permiso.estado_inicial,
        }
        for permiso in permisos
    ]

    return JsonResponse(permisos_list, safe=False)



def actualizar_estado_jefe(request):
    if request.method == "POST":
        permiso_id = request.POST.get("permiso_id")
        nuevo_estado = request.POST.get("nuevo_estado", "").strip().upper()  
        permiso_firmado = request.FILES.get("archivo_firmado")  # Recibe el archivo

        permiso = get_object_or_404(registroPermisos, id_permiso=permiso_id)

        if nuevo_estado == "APROBADO":
            permiso.estado_inicial = "PRE-APROBADO"
            permiso.descripcion = "APROBADO POR JEFE DE ÁREA"

            # Guardar el archivo firmado si se proporciona
            if permiso_firmado:
                permiso.permiso_firmado = permiso_firmado
                permiso.save()
                return JsonResponse({"status": "Success", "message": "Permiso aprobado correctamente."})

        elif nuevo_estado == "RECHAZADO":
            permiso.estado_inicial = "RECHAZADO"
            permiso.estado_final = "RECHAZADO"
            permiso.descripcion = "RECHAZADO POR JEFE DE ÁREA"
            permiso.save()
            return JsonResponse({"status": "Success", "message": "Permiso rechazado correctamente."})

        return JsonResponse({"status": "Error", "message": "Estado no válido."}, status=400)

    return JsonResponse({"status": "Error", "message": "Método no permitido."}, status=405)


def actualizar_estado_rrhh(request):
    if request.method == "POST":
        permiso_id = request.POST.get("permiso_id")
        nuevo_estado = request.POST.get("nuevo_estado")
        aprobado_por = request.user.username

        permiso = get_object_or_404(registroPermisos, id_permiso=permiso_id)

        if nuevo_estado == "APROBADO":
            permiso.estado_final = "APROBADO"
            permiso.descripcion = "APROBADO POR RRHH"
            permiso.modificado_por = aprobado_por
            permiso.fecha_modificacion = now()
            mensaje = "Permiso aprobado correctamente."
        elif nuevo_estado == "RECHAZADO":
            permiso.estado_final = "RECHAZADO"
            permiso.descripcion = "RECHAZADO POR RRHH"
            permiso.modificado_por = aprobado_por
            permiso.fecha_modificacion = now()
            mensaje = "Permiso rechazado correctamente."

        else:
            return JsonResponse({"status": "Error", "message": "Estado no válido."}, status=400)

        permiso.save()
        return JsonResponse({"status": "Success", "message": mensaje})

    return JsonResponse({"status": "Error", "message": "Método no permitido."}, status=405)

def colaboradores_con_permisos(request):
    """ Filtra los colaboradores con permisos en estado PENDIENTE o ACEPTADO por departamento """
    departamento_id = request.GET.get("departamento_id")
    nombre = request.GET.get("nombre", "").strip()

    if not departamento_id:
        return JsonResponse({"error": "Departamento no especificado."}, status=400)

    permisos = registroPermisos.objects.filter(
        id_departamento=departamento_id,
        estado_inicial__in=["PENDIENTE", "PRE-APROBADO"],
        comprobante=""
    ).select_related("codigocolaborador")

    if nombre:
        permisos = permisos.filter(codigocolaborador__nombrecolaborador__icontains=nombre)

    colaboradores = [
        {
            "id": permiso.id_permiso,  # Usamos el ID del permiso para guardar el comprobante
            "nombre": permiso.codigocolaborador.nombrecolaborador,
        }
        for permiso in permisos
    ]

    return JsonResponse({"colaboradores": colaboradores}, safe=False)
from django.views.decorators.csrf import csrf_exempt


def guardar_comprobante(request):
    """ Guarda el comprobante en la ruta definida en models.py """
    if request.method == "POST":
        permiso_id = request.POST.get("permiso_id")
        comprobante = request.FILES.get("comprobante")

        if not permiso_id or not comprobante:
            return JsonResponse({"error": "Datos incompletos."}, status=400)

        permiso = get_object_or_404(registroPermisos, id_permiso=permiso_id)
        permiso.comprobante = comprobante  
        permiso.save()

        return JsonResponse({"message": "Comprobante guardado exitosamente."})

    return JsonResponse({"error": "Método no permitido."}, status=405)
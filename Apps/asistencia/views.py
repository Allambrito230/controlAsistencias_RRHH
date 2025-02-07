# apps/asistencias/views.py

import logging
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import JsonResponse
from django.db.models import Q
from django.core.management import call_command
import io

from Apps.permisos.models import Colaboradores, Sucursal
from Apps.roles.models import Rol, RolAsignado
from .models import RegistroAsistencia
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, Color
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.utils.timezone import is_aware

logger = logging.getLogger(__name__)


@login_required
@permission_required('asistencia.view_registroasistencia', raise_exception=True)
def registroasistencia_list(request):
    """
    Vista que solo renderiza la plantilla de la tabla, 
    más el JSON de colaboradores (para autollenar en el modal "Crear").
    """
    colaboradores = Colaboradores.objects.filter(
        estado='ACTIVO').order_by("nombrecolaborador")
    colaborador_info = {}
    for c in colaboradores:
        sucursal_id = c.sucursal.id if c.sucursal else ""
        sucursal_nombre = c.sucursal.nombre_sucursal if c.sucursal else "N/A"

        def formatear_hhmm(value):
            """Devuelve hora en formato 'HH:MM' o 'N/A' si es None."""
            if value:
                return value.strftime("%H:%M")
            return "N/A"

        # Intentamos obtener un rol asignado activo (opcionalmente, podrías traer más de uno)
        rol_asignado = RolAsignado.objects.filter(
            colaborador=c, estado='ACTIVO').first()
        if rol_asignado and rol_asignado.rol:
            rol_id = rol_asignado.rol.id
            rol_nombre = rol_asignado.rol.nombre

            # Formateamos hora de inicio y fin para días de semana
            hora_inicio_semana = formatear_hhmm(
                rol_asignado.rol.hora_inicio_semana)
            hora_fin_semana = formatear_hhmm(rol_asignado.rol.hora_fin_semana)
            rol_horario_semana = f"{hora_inicio_semana} - {hora_fin_semana}"

            # Sábado
            hora_inicio_sabado = formatear_hhmm(
                rol_asignado.rol.hora_inicio_sabado)
            hora_fin_sabado = formatear_hhmm(rol_asignado.rol.hora_fin_sabado)
            rol_horario_sabado = f"{hora_inicio_sabado} - {hora_fin_sabado}"

            # Domingo
            hora_inicio_domingo = formatear_hhmm(
                rol_asignado.rol.hora_inicio_domingo)
            hora_fin_domingo = formatear_hhmm(
                rol_asignado.rol.hora_fin_domingo)
            rol_horario_domingo = f"{hora_inicio_domingo} - {hora_fin_domingo}"
        else:
            rol_id = ""
            rol_nombre = "Sin Rol"
            rol_horario_semana = "N/A"
            rol_horario_sabado = "N/A"
            rol_horario_domingo = "N/A"

        colaborador_info[c.id] = {
            # si quieres mostrar nombre en el modal de edición
            "colaborador_nombre": c.nombrecolaborador,
            "sucursal_id": sucursal_id,
            "sucursal_nombre": sucursal_nombre,
            "rol_id": rol_id,
            "rol_nombre": rol_nombre,
            "rol_horario_semana": rol_horario_semana,
            "rol_horario_sabado": rol_horario_sabado,
            "rol_horario_domingo": rol_horario_domingo
        }

    context = {
        'colaboradores': colaboradores,
        'colaborador_info': colaborador_info
    }
    return render(request, 'registroasistencia_crud.html', context)


@login_required
@permission_required('asistencia.view_registroasistencia', raise_exception=True)
def registroasistencia_list_api(request):
    """
    Endpoint AJAX para DataTables (Server-Side). Devuelve registros en JSON según paginación.
    """
    draw = request.POST.get('draw', 0)
    start = int(request.POST.get('start', 0))
    length = int(request.POST.get('length', 10))
    search_value = request.POST.get('search[value]', '')

    order_column_index = request.POST.get('order[0][column]', None)
    order_direction = request.POST.get('order[0][dir]', 'asc')

    columns = [
        'colaborador__nombrecolaborador',
        'colaborador__codigocolaborador',
        'sucursal__nombre_sucursal',
        'rol',         # interpretado dinamicamente
        'fecha',
        'hora_entrada',
        'hora_salida',
        'cumplimiento',
        'total_horas',
        'justificado',
        'estado',
        'id',          # para acciones
    ]

    queryset = (RegistroAsistencia.objects
                .filter(estado='ACTIVO')
                .select_related('colaborador', 'sucursal', 'rol'))

    # 1) Filtrado por search_value
    if search_value:
        queryset = queryset.filter(
            Q(colaborador__nombrecolaborador__icontains=search_value) |
            Q(colaborador__codigocolaborador__icontains=search_value) |
            Q(sucursal__nombre_sucursal__icontains=search_value) |
            Q(rol__nombre__icontains=search_value)
        )

    records_filtered = queryset.count()

    # 2) Ordenar
    if order_column_index is not None:
        try:
            order_column_index = int(order_column_index)
            order_column = columns[order_column_index]
            if order_direction == 'desc':
                order_column = '-' + order_column
            queryset = queryset.order_by(order_column)
        except (ValueError, IndexError):
            pass

    # 3) Paginación
    queryset = queryset[start:start+length]

    # 4) Construir data en formato [ [col0, col1, col2...], [...], ... ]
    data = []
    for reg in queryset:
        # Determinamos el horario según día de la semana
        if reg.rol:
            day_week = reg.fecha.weekday()  # 0=lunes,...,6=domingo
            if day_week == 6:
                horario_rol = f"{reg.rol.hora_inicio_domingo} - {reg.rol.hora_fin_domingo}"
            elif day_week == 5:
                horario_rol = f"{reg.rol.hora_inicio_sabado} - {reg.rol.hora_fin_sabado}"
            else:
                horario_rol = f"{reg.rol.hora_inicio_semana} - {reg.rol.hora_fin_semana}"
        else:
            horario_rol = "Sin Rol (Fuera de rango)"

        # Cumplimiento
        if reg.cumplimiento == "NO_CUMPLIO":
            cumplimiento_html = '<span class="badge text-bg-danger">No Cumplió</span>'
        elif reg.cumplimiento in ["<", ">", "="]:
            cumplimiento_html = f'<span class="badge text-bg-success">Cumplió | {reg.cumplimiento}</span>'
        else:
            cumplimiento_html = reg.cumplimiento

        # Justificado
        justificado_html = (
            '<span class="badge text-bg-warning">Con Justificación</span>'
            if reg.justificado else
            '<span class="badge text-bg-secondary">Sin Justificación</span>'
        )

        # Acciones
        acciones_html = f"""
        <div class="btn-group">
            <button class="btn btn-warning btn-sm btn-edit" data-id="{reg.id}">
              Editar
            </button>
            <button class="btn btn-danger btn-sm btn-inactivate" data-id="{reg.id}">
              Inactivar
            </button>
        </div>
        """

        data.append([
            reg.colaborador.nombrecolaborador,
            reg.colaborador.codigocolaborador,
            reg.sucursal.nombre_sucursal,
            horario_rol,
            reg.fecha.strftime("%Y-%m-%d"),
            reg.hora_entrada.strftime("%H:%M") if reg.hora_entrada else "",
            reg.hora_salida.strftime("%H:%M") if reg.hora_salida else "",
            cumplimiento_html,
            str(reg.total_horas) if reg.total_horas else "",
            justificado_html,
            reg.estado,
            acciones_html,
        ])

    response = {
        'draw': draw,
        'recordsTotal': RegistroAsistencia.objects.filter(estado='ACTIVO').count(),
        'recordsFiltered': records_filtered,
        'data': data,
    }
    return JsonResponse(response)


@login_required
@permission_required('asistencia.view_registroasistencia', raise_exception=True)
def registroasistencia_detail(request, registro_id):
    """
    Devuelve en JSON los datos de un RegistroAsistencia,
    para rellenar el modal 'Editar' via JavaScript.
    """
    try:
        reg = get_object_or_404(
            RegistroAsistencia, pk=registro_id, estado='ACTIVO')
        # Construimos un dict con los campos necesarios
        data = {
            'id': reg.id,
            'colaborador_id': reg.colaborador.id if reg.colaborador else None,
            'sucursal_id': reg.sucursal.id if reg.sucursal else None,
            'rol_id': reg.rol.id if reg.rol else None,
            'fecha': reg.fecha.strftime("%Y-%m-%d") if reg.fecha else "",
            'hora_entrada': reg.hora_entrada.strftime("%H:%M") if reg.hora_entrada else "",
            'hora_salida': reg.hora_salida.strftime("%H:%M") if reg.hora_salida else "",
        }
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def registroasistencia_create(request):
    """
    Lógica para crear un nuevo RegistroAsistencia (similar a lo que tenías).
    """
    if request.method == 'POST':
        try:
            # Aquí extraes los campos del request y creas el registro
            colaborador_id = request.POST.get('colaborador')
            sucursal_id = request.POST.get('sucursal')
            rol_id = request.POST.get('rol')
            fecha_str = request.POST.get('fecha')
            hora_entrada_str = request.POST.get('hora_entrada')
            hora_salida_str = request.POST.get('hora_salida')

            colaborador = get_object_or_404(Colaboradores, pk=colaborador_id)
            sucursal = get_object_or_404(
                Sucursal, pk=sucursal_id) if sucursal_id else None
            rol = get_object_or_404(Rol, pk=rol_id) if rol_id else None

            # Convierte a datetime/time
            from datetime import datetime

            def str_to_date(s):
                return datetime.strptime(s, "%Y-%m-%d").date() if s else None

            def str_to_time(s):
                return datetime.strptime(s, "%H:%M").time() if s else None

            fecha_obj = str_to_date(fecha_str)
            hora_entrada = str_to_time(
                hora_entrada_str) if hora_entrada_str else None
            hora_salida = str_to_time(
                hora_salida_str) if hora_salida_str else None

            RegistroAsistencia.objects.create(
                colaborador=colaborador,
                sucursal=sucursal,
                rol=rol,
                fecha=fecha_obj,
                hora_entrada=hora_entrada,
                hora_salida=hora_salida,
                creado_por=request.user
            )
            messages.success(
                request, "Registro de asistencia creado correctamente.")
        except Exception as e:
            logger.exception("Error al crear registro de asistencia")
            messages.error(
                request, f"Error al crear registro de asistencia: {e}")
    return redirect('registroasistencia_list')


@login_required
def registroasistencia_update(request, registro_id):
    """
    Lógica para actualizar un RegistroAsistencia (usada desde el modal 'Editar').
    """
    if request.method == 'POST':
        try:
            registro = get_object_or_404(
                RegistroAsistencia, pk=registro_id, estado='ACTIVO')

            colaborador_id = request.POST.get('colaborador')
            sucursal_id = request.POST.get('sucursal')
            rol_id = request.POST.get('rol')
            fecha_str = request.POST.get('fecha')
            hora_entrada_str = request.POST.get('hora_entrada')
            hora_salida_str = request.POST.get('hora_salida')

            colaborador = get_object_or_404(
                Colaboradores, pk=colaborador_id) if colaborador_id else None
            sucursal = get_object_or_404(
                Sucursal, pk=sucursal_id) if sucursal_id else None
            rol = get_object_or_404(Rol, pk=rol_id) if rol_id else None

            def str_to_date(s):
                return datetime.strptime(s, "%Y-%m-%d").date() if s else None

            def str_to_time(s):
                return datetime.strptime(s, "%H:%M").time() if s else None

            registro.colaborador = colaborador
            registro.sucursal = sucursal
            registro.rol = rol
            registro.fecha = str_to_date(fecha_str)
            registro.hora_entrada = str_to_time(
                hora_entrada_str) if hora_entrada_str else None
            registro.hora_salida = str_to_time(
                hora_salida_str) if hora_salida_str else None
            registro.modificado_por = request.user
            registro.save()
            messages.success(
                request, "Registro de asistencia actualizado correctamente.")
        except Exception as e:
            logger.exception(
                f"Error al actualizar registro de asistencia {registro_id}")
            messages.error(request, f"Error al actualizar: {e}")
    return redirect('registroasistencia_list')


@login_required
def registroasistencia_inactivate(request, registro_id):
    """
    Lógica para inactivar un RegistroAsistencia (botón "Inactivar").
    """
    if request.method == 'POST':
        try:
            registro = get_object_or_404(
                RegistroAsistencia, pk=registro_id, estado='ACTIVO')
            registro.estado = 'INACTIVO'
            registro.modificado_por = request.user
            registro.save()
            messages.success(
                request, "Registro de asistencia inactivado correctamente.")
        except Exception as e:
            logger.exception(
                f"Error al inactivar registro de asistencia {registro_id}")
            messages.error(request, f"Error al inactivar: {e}")
    return redirect('registroasistencia_list')


@login_required
@permission_required('asistencia.change_registroasistencia', raise_exception=True)
def sync_biometrico_view(request):
    """
    Llama al comando 'sync_biometrico' para sincronizar datos.
    """
    if request.method == 'POST':
        try:
            output = io.StringIO()
            call_command('sync_biometrico', stdout=output)
            messages.success(
                request, "Sincronización ejecutada exitosamente. Salida:\n" + output.getvalue())
            #return redirect('registroasistencia_list')
        except Exception as e:
            messages.error(request, f"Error al sincronizar: {str(e)}")
        return redirect('registroasistencia_list')
    else:
        return redirect('registroasistencia_list')


def exportar_asistencias_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro de Asistencias"

    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")

    thin_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )

    # Columnas deseadas (sin 'Estado', incluye 'Total Horas')
    #  1) ID
    #  2) Colaborador
    #  3) Código
    #  4) Sucursal
    #  5) Rol (horario según día)
    #  6) Fecha
    #  7) Hora Entrada
    #  8) Hora Salida
    #  9) Cumplimiento
    # 10) Justificado
    # 11) Total Horas
    columnas = [
        "ID",
        "Colaborador",
        "Código",
        "Sucursal",
        "Rol (Horario por Día)",
        "Fecha",
        "Hora Entrada",
        "Hora Salida",
        "Cumplimiento",
        "Justificado",
        "Total Horas",
    ]

    ws.append(columnas)

    # Estilos para la cabecera
    for col_num, col_name in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Helper para formatear horas HH:MM
    def formatear_hhmm(hora):
        if hora:
            return hora.strftime("%H:%M")
        return ""

    # Lógica para obtener la "franja horaria" (rol del día)
    def obtener_horario_dia(reg):
        """
        Retorna la cadena 'hh:mm - hh:mm' correspondiente a un día de semana, sábado o domingo,
        o 'Sin Rol (Fuera de rango)' si no hay rol asignado.
        """
        if not reg.rol:
            return "Sin Rol (Fuera de rango)"

        day_week = reg.fecha.weekday()  # 0 = lunes, ... 6 = domingo
        if day_week == 6:
            # domingo
            hi = formatear_hhmm(reg.rol.hora_inicio_domingo)
            hf = formatear_hhmm(reg.rol.hora_fin_domingo)
        elif day_week == 5:
            # sábado
            hi = formatear_hhmm(reg.rol.hora_inicio_sabado)
            hf = formatear_hhmm(reg.rol.hora_fin_sabado)
        else:
            # lunes-viernes
            hi = formatear_hhmm(reg.rol.hora_inicio_semana)
            hf = formatear_hhmm(reg.rol.hora_fin_semana)

        return f"{hi} - {hf}" if hi or hf else "Sin Rol (Fuera de rango)"

    # Query de registros (puedes filtrar si quieres solo los 'ACTIVO', etc.)
    queryset = (RegistroAsistencia.objects
                .select_related('colaborador', 'sucursal', 'rol')
                .all())

    row_index = 2
    for reg in queryset:
        # Preparamos datos
        rol_str = obtener_horario_dia(reg)

        fila = [
            reg.id,
            reg.colaborador.nombrecolaborador if reg.colaborador else "",
            reg.colaborador.codigocolaborador if reg.colaborador else "",
            reg.sucursal.nombre_sucursal if reg.sucursal else "",
            rol_str,
            reg.fecha,  # date
            formatear_hhmm(reg.hora_entrada),  # hora_entrada en HH:MM
            formatear_hhmm(reg.hora_salida),   # hora_salida en HH:MM
            reg.cumplimiento,  # 'NO_CUMPLIO', '<', '>', '='
            reg.justificado,   # bool
            str(reg.total_horas) if reg.total_horas else "",  # ej: "8:00:00"
        ]

        ws.append(fila)

        # Opcional: Eliminar tz
        # for date_idx in [6]:  # si reg.fecha fuera tz aware. (En tu caso es date, normal.)
        #     cell_dt = ws.cell(row=row_index, column=date_idx)
        #     if isinstance(cell_dt.value, (datetime.datetime, datetime.date)) and is_aware(cell_dt.value):
        #         cell_dt.value = cell_dt.value.replace(tzinfo=None)

        # Colorear "Cumplimiento" => 9a col
        cumplimiento_cell = ws.cell(row=row_index, column=9)
        if reg.cumplimiento == "NO_CUMPLIO":
            cumplimiento_cell.fill = PatternFill(
                start_color="DC3545", fill_type="solid")  # rojo
            cumplimiento_cell.font = Font(color="FFFFFF")
        elif reg.cumplimiento in ["<", ">", "="]:
            cumplimiento_cell.fill = PatternFill(
                start_color="28A745", fill_type="solid")  # verde
            cumplimiento_cell.font = Font(color="FFFFFF")
        else:
            # Sin relleno o algún otro
            pass

        # Colorear "Justificado" => 10a col
        justificado_cell = ws.cell(row=row_index, column=10)
        if reg.justificado is True:
            justificado_cell.fill = PatternFill(
                start_color="FFC107", fill_type="solid")  # amarillo
            justificado_cell.font = Font(color="000000", bold=True)
        else:
            justificado_cell.fill = PatternFill(
                start_color="6C757D", fill_type="solid")  # gris
            justificado_cell.font = Font(color="FFFFFF")

        # Bordes para toda la fila
        for col_num in range(1, len(columnas) + 1):
            c = ws.cell(row=row_index, column=col_num)
            c.border = thin_border

        row_index += 1

    # Crear la tabla con estilo "verde medio 21"
    ultima_fila = row_index - 1
    ultima_columna = len(columnas)
    from openpyxl.utils import get_column_letter
    rango_tabla = f"A1:{get_column_letter(ultima_columna)}{ultima_fila}"

    tabla = Table(displayName="TablaAsistencias", ref=rango_tabla)
    estilo_tabla = TableStyleInfo(
        name="TableStyleMedium21",  # Verde medio
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.tableStyleInfo = estilo_tabla
    ws.add_table(tabla)

    # Ajustar ancho de columnas
    for col_idx, _ in enumerate(columnas, start=1):
        max_length = 0
        letter = get_column_letter(col_idx)
        for cell in ws[letter]:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[letter].width = max_length + 2

    # Retornar como archivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="RegistroAsistencias.xlsx"'
    wb.save(response)
    return response
